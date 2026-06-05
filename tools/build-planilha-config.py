from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "central-configuracao-krisna-fernando.xlsx"

PALETTE = {
    "olive": "4E5A42",
    "sage": "75806B",
    "gold": "B29A68",
    "cream": "F8F4EC",
    "paper": "FFFAF2",
    "ink": "28241F",
    "muted": "756E62",
}

SHEETS = [
    {
        "name": "INSTRUCOES",
        "rows": [
            ["Central de configuração do site Krisna & Fernando"],
            ["Use esta planilha como base. Depois de importar para o Google Sheets, publique o Apps Script como App da Web e cole a URL em js/config.js."],
            [""],
            ["Abas principais"],
            ["CONFIG", "Configurações globais do site em chave/valor."],
            ["PAGINAS", "Textos e blocos editáveis por seção."],
            ["CONVIDADOS", "Base futura de convidados e controle interno."],
            ["RSVP", "Respostas enviadas pelo formulário do site."],
            ["PRESENTES", "Lista, PIX, cotas ou links de presentes."],
            ["GALERIA", "Fotos que podem ser exibidas no site."],
            ["MENSAGENS", "Mensagens ou depoimentos opcionais."],
            [""],
            ["Publicação do Apps Script"],
            ["1. Extensões > Apps Script"],
            ["2. Cole o conteúdo de appscript/Code.gs"],
            ["3. Implantar > Nova implantação > App da Web"],
            ["4. Executar como: Eu"],
            ["5. Quem pode acessar: Qualquer pessoa"],
            ["6. Copie a URL gerada e cole em js/config.js"],
        ],
        "widths": [36, 100],
    },
    {
        "name": "CONFIG",
        "rows": [
            ["key", "value", "description"],
            ["couple_name", "Krisna & Fernando", "Nome exibido no site"],
            ["domain", "https://krisnaefernando.com/", "Domínio final"],
            ["wedding_date", "2026-10-29T15:30:00-03:00", "Data/hora da contagem regressiva"],
            ["ceremony_date_label", "29 de outubro de 2026", "Texto amigável da data"],
            ["ceremony_time", "15h30", "Horário da cerimônia"],
            ["venue_name", "Buffet La Maison", "Nome do local"],
            ["ceremony_room", "Salão Terrasse", "Sala da cerimônia"],
            ["reception_room", "Salão Central", "Sala da recepção"],
            ["venue_address", "Av. Eng. Luiz Vieira, 555, Papicu", "Endereço"],
            ["maps_url", "https://www.google.com/maps/search/?api=1&query=Buffet%20La%20Maison%20Av.%20Eng.%20Luiz%20Vieira%20555%20Papicu", "Link de rota"],
            ["whatsapp_contact", "", "Contato do casal ou assessoria"],
            ["instagram_url", "", "Instagram opcional"],
            ["rsvp_enabled", "TRUE", "Ativa/desativa formulário"],
            ["gifts_enabled", "TRUE", "Ativa/desativa presentes"],
            ["intro_enabled", "TRUE", "Ativa/desativa intro em vídeo"],
            ["site_status", "preview", "preview ou published"],
        ],
        "widths": [28, 78, 56],
    },
    {
        "name": "PAGINAS",
        "rows": [
            ["section", "title", "subtitle", "text", "button_label", "button_url", "enabled", "sort_order"],
            ["hero", "Krisna & Fernando", "29 de outubro de 2026", "Com amor, presença de Deus e a alegria de quem encontrou no outro o seu lugar.", "Confirmar presença", "#presenca", "TRUE", 1],
            ["historia", "Uma estampa feita de encontros, fé e detalhes nossos.", "Nossa história preferida", "Nossa estampa foi criada especialmente para o casamento, unindo elementos que fazem parte da nossa história.", "", "", "TRUE", 2],
            ["casamento", "Esperamos vocês para celebrar esse dia conosco.", "O casamento", "A cerimônia acontecerá às 15h30 do dia 29 de outubro de 2026, no Buffet La Maison, no Salão Terrasse. Após a cerimônia, os convidados serão recepcionados no mesmo local, no Salão Central.", "Abrir rota", "{{maps_url}}", "TRUE", 3],
            ["presenca", "Estamos preparando tudo com muito amor e presença de Deus.", "Confirme sua presença", "Para que possamos organizar da melhor forma possível esse dia tão especial, pedimos que nos confirme sua presença.", "Enviar confirmação", "", "TRUE", 4],
            ["presentes", "Seu carinho já é parte da nossa casa.", "Presentes", "Em breve, esta seção receberá nossa lista de presentes, cotas e informações configuradas por aqui.", "", "", "TRUE", 5],
            ["rodape", "Krisna & Fernando", "", "Obrigado por fazer parte da nossa história.", "", "", "TRUE", 6],
        ],
        "widths": [18, 46, 30, 92, 24, 46, 14, 14],
    },
    {
        "name": "CONVIDADOS",
        "rows": [
            ["guest_id", "name", "phone", "email", "group", "allowed_companions", "notes", "status"],
            ["KF-001", "Convidado Exemplo", "", "", "Família", 1, "Substituir por convidados reais", "pendente"],
        ],
        "widths": [16, 34, 22, 34, 20, 22, 44, 18],
    },
    {
        "name": "RSVP",
        "rows": [
            ["timestamp", "name", "phone", "email", "companion", "source", "userAgent"],
            ["", "", "", "", "", "", ""],
        ],
        "widths": [24, 34, 22, 34, 34, 16, 70],
    },
    {
        "name": "PRESENTES",
        "rows": [
            ["gift_id", "title", "description", "type", "url", "pix_key", "amount", "enabled", "sort_order"],
            ["P-001", "Lista de presentes", "Link provisório para lista externa.", "link", "", "", "", "FALSE", 1],
            ["P-002", "Cota lua de mel", "Cota simbólica para a viagem dos noivos.", "cota", "", "", 250, "FALSE", 2],
            ["P-003", "PIX dos noivos", "Informação provisória para contribuição via PIX.", "pix", "", "", "", "FALSE", 3],
        ],
        "widths": [14, 30, 58, 18, 48, 34, 16, 14, 14],
    },
    {
        "name": "GALERIA",
        "rows": [
            ["image_id", "file_name", "title", "alt_text", "featured", "enabled", "sort_order"],
            ["G-001", "_BQH1996.jpg", "Ensaio Krisna e Fernando", "Krisna e Fernando em ensaio do casal", "TRUE", "TRUE", 1],
            ["G-002", "_BQH1929-Editar.jpg", "Ensaio do casal", "Foto do casal em ensaio pré-casamento", "TRUE", "TRUE", 2],
            ["G-003", "_BQH1901.jpg", "Ensaio do casal", "Foto de Krisna e Fernando", "FALSE", "TRUE", 3],
            ["G-004", "_BQH1920.jpg", "Ensaio do casal", "Foto de Krisna e Fernando", "FALSE", "TRUE", 4],
        ],
        "widths": [14, 28, 34, 54, 16, 14, 14],
    },
    {
        "name": "MENSAGENS",
        "rows": [
            ["message_id", "name", "message", "enabled", "sort_order"],
            ["M-001", "Krisna & Fernando", "Obrigado por fazer parte desse capítulo tão especial da nossa história.", "TRUE", 1],
        ],
        "widths": [16, 28, 86, 14, 14],
    },
]


def style_sheet(ws, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    max_col = ws.max_column
    header_fill = PatternFill("solid", fgColor=PALETTE["olive"])
    body_fill = PatternFill("solid", fgColor=PALETTE["paper"])

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = body_fill
            cell.font = Font(color=PALETTE["ink"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for index in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(index)].width = widths[index - 1] if index <= len(widths) else 22

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def build():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for spec in SHEETS:
        ws = wb.create_sheet(spec["name"])
        for row in spec["rows"]:
            ws.append(row)
        style_sheet(ws, spec["widths"])

        if spec["name"] == "INSTRUCOES":
            ws.merge_cells("A1:B1")
            ws["A1"].fill = PatternFill("solid", fgColor=PALETTE["olive"])
            ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
            ws.merge_cells("A2:B2")
            ws["A2"].fill = PatternFill("solid", fgColor=PALETTE["cream"])
            ws["A2"].font = Font(color=PALETTE["muted"])
            ws.freeze_panes = None
            ws.auto_filter.ref = None

    wb.save(OUTPUT)

    # Open once to verify the workbook is readable.
    check = load_workbook(OUTPUT, read_only=True)
    expected = [sheet["name"] for sheet in SHEETS]
    if check.sheetnames != expected:
        raise RuntimeError(f"Abas inesperadas: {check.sheetnames}")
    check.close()
    print(OUTPUT)


if __name__ == "__main__":
    build()
